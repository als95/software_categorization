import openpyxl
import time
from datetime import datetime
from selenium import webdriver

wb = openpyxl.load_workbook('E:\\data\\label\\category_label.xlsx')
sheet = wb['Sheet1']

rows = sheet.rows # need decision
cols = 8

for row in range(len(rows)):
    for col in range(len(cols)):
        if sheet.cell(row=row+1, column=col+1).value is None:
            sheet.cell(row+1, col+1, "dummy")